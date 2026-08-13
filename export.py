# -*- coding: utf-8 -*-
"""
认证号提取 + Excel 导出模块
=============================

功能：
  1. 从认证 PDF 中提取认证号（Reference No / Report No 等）
  2. 批量提取所有产品的认证号，存入数据库
  3. 导出 Excel 认证映射表

用法：
  python export.py --extract-pdfs       # 批量提取所有认证 PDF 的认证号
  python export.py --excel             # 导出 Excel 认证映射表
  python export.py --excel --output /path/to/cert.xlsx  # 指定输出路径
"""

import argparse
import io
import json
import os
import re
import sqlite3
import sys
import time
from datetime import datetime, timezone

import requests
from pypdf import PdfReader
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from scraper import (
    init_db, DB_PATH, HEADERS, CERT_TYPE_NAMES,
    get_product, save_product, fetch_sitemap_products,
)

PDF_TIMEOUT = 20
PDF_DELAY = 0.5

# 认证号提取规则：按优先级排列
CERT_NUMBER_PATTERNS = [
    (r"[Rr]eference\s*(?:No|Number|#)[:\s.#]*([A-Z0-9][-A-Z0-9/.]+)", "Reference No"),
    (r"[Rr]eport\s*(?:No|Number|#)[:\s.#]*([A-Z0-9][-A-Z0-9/.]+)", "Report No"),
    (r"[Cc]ertificate\s*(?:No|Number|#)[:\s.#]*([A-Z0-9][-A-Z0-9/.]+)", "Certificate No"),
    (r"[Rr]egistration\s*(?:No|Number|#)[:\s.#]*([A-Z0-9][-A-Z0-9/.]+)", "Registration No"),
    (r"[Dd]ocument\s*(?:No|Number|#)[:\s.#]*([A-Z0-9][-A-Z0-9/.]+)", "Document No"),
    (r"FCC\s*ID[:\s]*([A-Z0-9][-A-Z0-9]+)", "FCC ID"),
    (r"[Mm]odel\s*(?:No|Number|#)[:\s.#]*([A-Za-z0-9][-A-Za-z0-9._]+)", "Model No"),
]


def extract_pdf_text(pdf_url):
    """下载 PDF 并提取文本，返回 (text, is_image_pdf)"""
    try:
        resp = requests.get(pdf_url, headers=HEADERS, timeout=PDF_TIMEOUT)
        if resp.status_code != 200:
            return "", False
        pdf_data = resp.content
    except Exception:
        return "", False

    text = ""
    try:
        reader = PdfReader(io.BytesIO(pdf_data))
        for page in reader.pages:
            text += page.extract_text() or ""
    except Exception:
        pass

    if text.strip():
        return text, False

    # pypdf 提取不到文字 -> 图片型 PDF（扫描件）
    return "", True


def extract_cert_number(text, cert_type=""):
    """从 PDF 文本中提取认证号"""
    if not text or not text.strip():
        return ""

    for pattern, label in CERT_NUMBER_PATTERNS:
        matches = re.findall(pattern, text)
        if matches:
            for m in matches:
                if len(m) >= 5 and m.upper() != "NA":
                    return m.strip()
    return ""


def extract_all_cert_numbers(force=False):
    """批量提取所有产品的认证 PDF 中的认证号"""
    conn = init_db()
    rows = conn.execute(
        "SELECT sku, name, certifications FROM products WHERE certifications != '[]'"
    ).fetchall()
    total = len(rows)
    print(f"共 {total} 个产品有认证信息，开始提取认证号...")
    print(f"（每个 PDF 间隔 {PDF_DELAY} 秒）\n")

    updated, skipped, failed = 0, 0, 0

    for i, (sku, name, certs_json) in enumerate(rows, 1):
        certs = json.loads(certs_json) if certs_json else []
        changed = False

        for cert in certs:
            if cert.get("source") != "certificate":
                continue
            if cert.get("cert_number") and not force:
                continue

            pdf_url = cert.get("pdf_url", "")
            if not pdf_url:
                continue

            try:
                text, is_image = extract_pdf_text(pdf_url)
                if is_image:
                    cert["cert_number"] = "图片型PDF（需手动查看）"
                    cert["is_image_pdf"] = True
                    changed = True
                elif text:
                    number = extract_cert_number(text, cert.get("type", ""))
                    cert["cert_number"] = number if number else "未找到认证号"
                    changed = True
                else:
                    cert["cert_number"] = "PDF下载失败"
                    changed = True

                time.sleep(PDF_DELAY)
            except Exception as e:
                cert["cert_number"] = f"提取失败: {e}"
                changed = True
                failed += 1

        if changed:
            conn.execute(
                "UPDATE products SET certifications = ?, last_updated = ? WHERE sku = ?",
                (json.dumps(certs, ensure_ascii=False),
                 datetime.now(timezone.utc).isoformat(), sku),
            )
            conn.commit()
            updated += 1

        if i % 20 == 0 or i <= 3:
            cert_info = "; ".join(
                f"{c['type']}={c.get('cert_number', '?')}" for c in certs if c.get("cert_number")
            )
            print(f"  [{i}/{total}] SKU:{sku} | {name[:30]} | {cert_info}")

    conn.close()
    print(f"\n提取完成！更新:{updated} 跳过:{skipped} 失败:{failed} 总计:{total}")


def generate_excel(output_path=None):
    """生成 Excel 认证映射表"""
    if output_path is None:
        output_path = os.path.join(
            os.path.dirname(DB_PATH),
            f"认证信息映射表_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        )

    conn = init_db()
    rows = conn.execute(
        "SELECT sku, product_id, name, url, cert_text, certifications, datasheet_url, last_updated "
        "FROM products ORDER BY sku"
    ).fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "认证信息映射表"

    header_font = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="0066FF", end_color="0066FF", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_font = Font(name="微软雅黑", size=10)
    cell_align = Alignment(vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    headers = ["SKU", "产品名称", "认证类型", "认证号（PDF内）", "认证PDF链接", "产品页面(Landing Page)", "数据更新时间"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    ws.freeze_panes = "A2"

    row_num = 2
    for sku, pid, name, url, cert_text, certs_json, datasheet_url, updated in rows:
        certs = json.loads(certs_json) if certs_json else []

        if not certs:
            ws.cell(row=row_num, column=1, value=sku).font = cell_font
            ws.cell(row=row_num, column=2, value=name or "").font = cell_font
            ws.cell(row=row_num, column=3, value="NA").font = cell_font
            ws.cell(row=row_num, column=4, value="NA").font = cell_font
            ws.cell(row=row_num, column=5, value="NA").font = cell_font
            ws.cell(row=row_num, column=6, value=url or "").font = cell_font
            ws.cell(row=row_num, column=7, value=updated or "").font = cell_font
            for col in range(1, 8):
                ws.cell(row=row_num, column=col).alignment = cell_align
                ws.cell(row=row_num, column=col).border = thin_border
            row_num += 1
        else:
            for cert in certs:
                cert_type = cert.get("type", "")
                cert_name = CERT_TYPE_NAMES.get(cert_type, cert_type)
                cert_number = cert.get("cert_number", "")
                if not cert_number:
                    cert_number = "未提取（运行 export.py --extract-pdfs）"
                pdf_url = cert.get("pdf_url", "")

                ws.cell(row=row_num, column=1, value=sku).font = cell_font
                ws.cell(row=row_num, column=2, value=name or "").font = cell_font
                ws.cell(row=row_num, column=3, value=cert_name).font = cell_font
                ws.cell(row=row_num, column=4, value=cert_number).font = cell_font
                ws.cell(row=row_num, column=5, value=pdf_url).font = cell_font
                ws.cell(row=row_num, column=6, value=url or "").font = cell_font
                ws.cell(row=row_num, column=7, value=updated or "").font = cell_font
                for col in range(1, 8):
                    ws.cell(row=row_num, column=col).alignment = cell_align
                    ws.cell(row=row_num, column=col).border = thin_border
                row_num += 1

    col_widths = [16, 40, 25, 30, 55, 55, 22]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 30

    wb.save(output_path)
    total_products = len(rows)
    total_rows = row_num - 2
    print(f"\nExcel 已生成: {output_path}")
    print(f"共 {total_products} 个产品，{total_rows} 行数据")
    return output_path


def main():
    parser = argparse.ArgumentParser(description="认证号提取 + Excel 导出")
    parser.add_argument("--extract-pdfs", action="store_true", help="批量提取所有认证 PDF 的认证号")
    parser.add_argument("--excel", action="store_true", help="导出 Excel 认证映射表")
    parser.add_argument("--output", type=str, default=None, help="Excel 输出路径")
    parser.add_argument("--force", action="store_true", help="强制重新提取")
    args = parser.parse_args()

    if args.extract_pdfs:
        print("=" * 55)
        print("  批量提取认证 PDF 中的认证号")
        print("=" * 55)
        extract_all_cert_numbers(force=args.force)
    elif args.excel:
        print("=" * 55)
        print("  导出 Excel 认证映射表")
        print("=" * 55)
        generate_excel(args.output)
    else:
        parser.print_help()


if __name__ == "__main__":
    main()
